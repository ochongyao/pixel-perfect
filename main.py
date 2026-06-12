"""Image ↔ Excel pixel art converter."""

import argparse
import logging
import sys
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(message)s")
log = logging.getLogger(__name__)

CELL_W = 3.0
CELL_H = 18.0
WHITE = 255, 255, 255
MAX_EXCEL_ROWS = 1_048_576
MAX_EXCEL_COLS = 16_384
CELL_WARN = 500_000


def image_to_excel(image_path, excel_path, scale=1.0):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    from PIL import Image, UnidentifiedImageError

    try:
        img = Image.open(image_path)
    except FileNotFoundError:
        sys.exit(f"Error: image not found — {image_path}")
    except UnidentifiedImageError:
        sys.exit(f"Error: cannot decode image — {image_path}")

    if scale <= 0:
        sys.exit(f"Error: scale must be positive, got {scale}")

    grid_width = max(1, round(img.width * scale))
    grid_height = max(1, round(img.height * scale))

    if grid_width > MAX_EXCEL_COLS or grid_height > MAX_EXCEL_ROWS:
        sys.exit(
            f"Error: grid {grid_width}×{grid_height} exceeds Excel limits "
            f"({MAX_EXCEL_COLS} cols × {MAX_EXCEL_ROWS} rows)"
        )

    cells = grid_width * grid_height
    if cells > CELL_WARN:
        log.warning("%d cells — Excel may be slow or crash", cells)

    log.info("→ %s  grid=%dx%d", image_path, grid_width, grid_height)
    pixels = list(
        img.resize((grid_width, grid_height), Image.Resampling.NEAREST)
        .convert("RGB")
        .getdata()
    )
    img.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Pixel Art"
    fill_cache = {}
    for i, (r, g, b) in enumerate(pixels):
        h = f"{r:02x}{g:02x}{b:02x}"
        if h not in fill_cache:
            fill_cache[h] = PatternFill(start_color=h, end_color=h, fill_type="solid")
        ws.cell(row=i // grid_width + 1, column=i % grid_width + 1).fill = fill_cache[h]
    for c in range(1, grid_width + 1):
        ws.column_dimensions[get_column_letter(c)].width = CELL_W
    for r in range(1, grid_height + 1):
        ws.row_dimensions[r].height = CELL_H

    try:
        wb.save(excel_path)
    except OSError as e:
        sys.exit(f"Error: cannot save workbook — {e}")
    log.info("  saved %s  (%d colours)", excel_path, len(fill_cache))


def _cell_colour(cell):
    try:
        rgb = cell.fill.start_color.rgb
        if not rgb:
            return WHITE
        if len(rgb) == 6:
            return int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
        if len(rgb) >= 8:
            return int(rgb[2:4], 16), int(rgb[4:6], 16), int(rgb[6:8], 16)
        return WHITE
    except (AttributeError, ValueError, TypeError):
        return WHITE


def excel_to_image(excel_path, output_image_path, scale=1):
    from openpyxl import load_workbook
    from PIL import Image

    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        sys.exit(f"Error: Excel file not found — {excel_path}")
    except Exception as e:
        sys.exit(f"Error: cannot read workbook — {e}")

    ws = wb.active
    if ws is None:
        sys.exit(f"Error: workbook has no sheets — {excel_path}")

    w, h = ws.max_column or 0, ws.max_row or 0
    if w < 1 or h < 1:
        sys.exit(f"Error: workbook appears empty — {excel_path}")

    if scale < 1:
        sys.exit(f"Error: scale must be at least 1, got {scale}")

    log.info("← %s  %dx%d  scale=×%d", excel_path, w, h, scale)
    img = Image.new("RGB", (w, h), color=WHITE)
    px = img.load()
    for r, row in enumerate(ws.iter_rows()):
        for c, cell in enumerate(row):
            px[c, r] = _cell_colour(cell)
    wb.close()

    if scale > 1:
        img = img.resize((w * scale, h * scale), Image.Resampling.NEAREST)

    try:
        img.save(output_image_path)
    except (OSError, ValueError) as e:
        sys.exit(f"Error: cannot save image — {e}")
    log.info("  saved %s", output_image_path)


def main():
    parser = argparse.ArgumentParser(
        description="Convert images to/from Excel pixel art."
    )
    sub = parser.add_subparsers(dest="command", required=True)

    to_excel = sub.add_parser(
        "to-excel", help="Convert an image to an Excel pixel grid"
    )
    to_excel.add_argument("image", type=Path)
    to_excel.add_argument("xlsx", type=Path)
    to_excel.add_argument("--scale", type=float, default=1.0,
                          help="Scale factor (default: 1.0)")

    to_image = sub.add_parser(
        "to-image", help="Convert an Excel pixel grid back to an image"
    )
    to_image.add_argument("xlsx", type=Path)
    to_image.add_argument("output", type=Path)
    to_image.add_argument("--scale", type=int, default=1)

    args = parser.parse_args()

    if args.command == "to-excel":
        image_to_excel(args.image, args.xlsx, args.scale)
    elif args.command == "to-image":
        excel_to_image(args.xlsx, args.output, args.scale)


if __name__ == "__main__":
    main()
